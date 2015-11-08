_author__ = 'Jacob Verghese, Sravanika Doddi'
#remove 'the', 'a', 'an' if the title string starts with it
import openpyxl
import urllib.request
from urllib.request import urlopen
import json

# GLOBAL_CONSTANTS - MODIFY ON TARGET MACHINE
spreadsheet_path = '/home/jacob/Desktop/augmented_data.xlsx'
omdb_url_path = 'http://www.omdbapi.com/?'

# RUNTIME_CONSTANTS
worksheet = openpyxl.load_workbook(spreadsheet_path)
movie_data = worksheet.get_sheet_by_name('Sheet1')
row_boundary = movie_data.max_row


#Currently somewhat slow, should reform with http.client
def get_movie_data(movie_title, year=''):
    """
    :param movie_title: the movie to query the OMDB API for.
    :param year: Defaults to an empty string when unspecified, else adds the year to the query.
    """
    user_agent = 'Mozilla/5.0 (X11; Linux i586; rv:31.0) Gecko/20100101 Firefox/31.0'
    headers = {'User-Agent': user_agent, }
    url = omdb_url_path + 't=' + movie_title + '&y=' + year + '&tomatoes=true'
    request = urllib.request.Request(url, None, headers)
    text = urlopen(request).read()
    return str(text)


def plusify(string):
    result = ''
    for character in string:
        if str(character).isalnum():
            result += character
        else:
            result += '+'

    return result



def deyearify(movie_title):
    """
    :param movie_title: Movie title containing the year of release to be removed (year is a separate field in query)
    :return: tuple of the movie name with year removed, and the year separately
    """
    index = movie_title.rfind('(')
    year = ''
    if index != -1:
        year = movie_title[(index + 1):(index + 5)]
        movie_title = movie_title[:(index - 1)]
        movie_title = '+'.join(movie_title.split())
    return str(movie_title), str(year)


def despacify(string):
    result = ''
    for character in string:
        if character != ' ':
            result += character
    return result

def deslashify(string):
    result = ''
    for character in string:
        result += character
        if character == '\\':
            result += character
    return result

def numerify(string):
    result = ''
    for character in string:
        if str(character).isdigit():
            result += character
    return int(character)


workbook = openpyxl.Workbook()
sheet = workbook.active
for row in range(2, row_boundary + 1):
    try:
        movie_name = str(movie_data['B' + str(row)].value)
        movie_name = deyearify(movie_name)
        string = get_movie_data(plusify(movie_name[0]), movie_name[1])
        string = deslashify(string[2:-1])
        data = json.loads(string)

        imdb_votes = str(data['imdbVotes'])
        sheet[str('P' + str(row))] = imdb_votes

        mpaa_rating = str(data['Rated'])
        sheet[str('Q' + str(row))] = mpaa_rating

        language = str(data['Language'])
        sheet[str('R' + str(row))] = language

        country = str(data['Country'])
        sheet[str('S' + str(row))] = country

        awards = str(data['Awards'])
        sheet[str('T' + str(row))] = awards

        imdb_id = str(data['imdbID'])
        sheet[str('U' + str(row))] = imdb_id

        sheet[str('A' + str(row))] = data['tomatoMeter']
        sheet[str('B' + str(row))] = data['tomatoImage']
        sheet[str('C' + str(row))] = data['tomatoReviews']
        sheet[str('D' + str(row))] = data['tomatoFresh']
        sheet[str('E' + str(row))] = data['tomatoRotten']
        sheet[str('F' + str(row))] = data['tomatoConsensus']
        sheet[str('G' + str(row))] = data['tomatoUserMeter']
        sheet[str('H' + str(row))] = data['tomatoUserRating']
        sheet[str('I' + str(row))] = data['tomatoUserReviews']

        print(row)
        workbook.save("help.xlsx")
    except:
        print("pass")
        pass



'''


        imdb_votes = data['imdbVotes']
        sheet[str('P' + str(row))] = imdb_votes

        mpaa_rating = data['Rated']
        sheet[str('Q' + str(row))] = mpaa_rating

        language = data['Language']
        sheet[str('R' + str(row))] = language

        country = data['Country']
        sheet[str('S' + str(row))] = country

        awards = data['Awards']
        sheet[str('T' + str(row))] = awards

        imdb_id = data['imdbID']
        sheet[str('U' + str(row))] = imdb_id

        sheet[str('A' + str(row))] = data['tomatoMeter']
        sheet[str('B' + str(row))] = data['tomatoImage']
        sheet[str('C' + str(row))] = data['tomatoReviews']
        sheet[str('D' + str(row))] = data['tomatoFresh']
        sheet[str('E' + str(row))] = data['tomatoRotten']
        sheet[str('F' + str(row))] = data['tomatoConsensus']
        sheet[str('G' + str(row))] = data['tomatoUserMeter']
        sheet[str('H' + str(row))] = data['tomatoUserRating']
        sheet[str('I' + str(row))] = data['tomatoUserReviews']

        print(row)
'''